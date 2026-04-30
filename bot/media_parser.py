"""Извлечение текста из медиафайлов для ingest-пайплайна.

Поддерживаемые форматы:
- PDF         — pypdf (pure python, нет системных зависимостей)
- DOCX/ODT    — python-docx / odfpy (опц.)
- XLSX/ODS    — openpyxl (опц.)
- TXT/MD/CSV  — прямо
- Голосовые   — faster-whisper (STT, модель tiny)
- Изображения — DeepSeek VL через describe_image

Все зависимости опциональны — при отсутствии возвращается заглушка.
"""
from __future__ import annotations

import asyncio
import base64
import logging
import tempfile
from pathlib import Path

log = logging.getLogger(__name__)

# ── PDF ──────────────────────────────────────────────────────────────────────

def _extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader  # type: ignore[import]
    except ImportError:
        return "(pypdf не установлен — pip install pypdf)"
    try:
        reader = PdfReader(str(path))
        pages: list[str] = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            if text.strip():
                pages.append(f"[стр. {i + 1}]\n{text.strip()}")
        if not pages:
            return "(PDF без извлекаемого текста — возможно, сканированный)"
        total = sum(len(p) for p in pages)
        result = "\n\n".join(pages)
        if total > 8000:
            result = result[:8000] + f"\n...(обрезано, всего ~{total} символов)"
        return result
    except Exception as exc:  # noqa: BLE001
        log.warning("PDF extract failed: %s", exc)
        return f"(ошибка чтения PDF: {exc})"


# ── DOCX ─────────────────────────────────────────────────────────────────────

def _extract_docx(path: Path) -> str:
    try:
        import docx  # type: ignore[import]
    except ImportError:
        return "(python-docx не установлен — pip install python-docx)"
    try:
        doc = docx.Document(str(path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n".join(paragraphs)
        if len(text) > 8000:
            text = text[:8000] + "\n...(обрезано)"
        return text or "(документ пуст)"
    except Exception as exc:  # noqa: BLE001
        log.warning("DOCX extract failed: %s", exc)
        return f"(ошибка чтения DOCX: {exc})"


# ── XLSX ─────────────────────────────────────────────────────────────────────

def _extract_xlsx(path: Path) -> str:
    try:
        import openpyxl  # type: ignore[import]
    except ImportError:
        return "(openpyxl не установлен — pip install openpyxl)"
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            lines.append(f"## Лист: {sheet}")
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    lines.append("\t".join(cells))
                    row_count += 1
                    if row_count >= 200:
                        lines.append("...(строки обрезаны)")
                        break
        text = "\n".join(lines)
        if len(text) > 8000:
            text = text[:8000] + "\n...(обрезано)"
        return text or "(таблица пуста)"
    except Exception as exc:  # noqa: BLE001
        log.warning("XLSX extract failed: %s", exc)
        return f"(ошибка чтения XLSX: {exc})"


# ── PLAIN TEXT ────────────────────────────────────────────────────────────────

def _extract_text(path: Path) -> str:
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
        if len(text) > 8000:
            text = text[:8000] + f"\n...(обрезано, всего {len(text)} символов)"
        return text
    except Exception as exc:  # noqa: BLE001
        return f"(ошибка чтения файла: {exc})"


# ── VOICE / AUDIO STT ─────────────────────────────────────────────────────────

def _transcribe_audio_sync(path: Path) -> str:
    """Синхронная транскрипция через openai-whisper (модель tiny, ffmpeg-based)."""
    try:
        import whisper  # type: ignore[import]
    except ImportError:
        return "(openai-whisper не установлен — pip install openai-whisper)"
    try:
        model = whisper.load_model("tiny")
        result = model.transcribe(str(path))
        text = (result.get("text") or "").strip()
        lang = result.get("language", "?")
        log.info("STT %s: lang=%s text=%r", path.name, lang, text[:80])
        return text or "(пустая транскрипция)"
    except Exception as exc:  # noqa: BLE001
        log.warning("STT failed for %s: %s", path, exc)
        return f"(ошибка STT: {exc})"


async def transcribe_audio(path: Path) -> str:
    """Асинхронная обёртка для STT (запускает в thread-pool)."""
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, _transcribe_audio_sync, path)


# ── MAIN ENTRY ────────────────────────────────────────────────────────────────

TEXT_EXTENSIONS = {".txt", ".md", ".csv", ".json", ".xml", ".html", ".py",
                   ".log", ".yaml", ".yml", ".toml", ".ini", ".cfg"}
AUDIO_EXTENSIONS = {".ogg", ".mp3", ".wav", ".m4a", ".flac", ".opus"}
VIDEO_EXTENSIONS = {".mp4", ".mov", ".avi", ".mkv", ".webm", ".3gp", ".ts"}


# ── VIDEO: STT + keyframes ────────────────────────────────────────────────────

async def _run_ffmpeg(args: list[str], timeout: float = 60.0) -> tuple[int, bytes, bytes]:
    """Запустить ffmpeg асинхронно с таймаутом; убивает процесс при превышении."""
    proc = await asyncio.create_subprocess_exec(
        *args,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
        start_new_session=True,  # отдельная группа — не получает SIGSTOP от job-control
    )
    try:
        stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
        return proc.returncode or 0, stdout, stderr
    except asyncio.TimeoutError:
        log.warning("ffmpeg timeout (%.0fs), killing pid=%s: %s", timeout, proc.pid, args[:4])
        try:
            proc.kill()
        except Exception:
            pass
        try:
            await asyncio.wait_for(proc.communicate(), timeout=5.0)
        except Exception:
            pass
        return -1, b"", b""


_VIDEO_AUDIO_SIZE_LIMIT = 50 * 1024 * 1024  # 50 MB — выше этого STT скипаем


async def _extract_video_audio_async(video_path: Path, audio_path: Path) -> bool:
    """Извлечь аудио из видео через ffmpeg в WAV (16 kHz mono)."""
    size = video_path.stat().st_size
    log.info("video audio extract start: %s (%.1fMB)",
             video_path.name, size / 1024 / 1024)
    if size > _VIDEO_AUDIO_SIZE_LIMIT:
        log.info("video too large for STT (%.1fMB > 50MB), skipping: %s",
                 size / 1024 / 1024, video_path.name)
        return False
    rc, _, stderr = await _run_ffmpeg(
        ["ffmpeg", "-y", "-i", str(video_path),
         "-vn", "-ar", "16000", "-ac", "1", "-f", "wav", str(audio_path)],
        timeout=120.0,
    )
    if rc != 0:
        log.info("video audio extract failed rc=%d: %s", rc, stderr[-200:] if stderr else b"")
    else:
        sz = audio_path.stat().st_size if audio_path.exists() else 0
        log.info("video audio extract ok: %s -> %dKB wav", video_path.name, sz // 1024)
    return rc == 0


async def _extract_keyframes_async(video_path: Path, frames_dir: Path, n: int = 4) -> list[Path]:
    """Извлечь N равномерно распределённых кейфреймов из видео через ffmpeg."""
    frames_dir.mkdir(parents=True, exist_ok=True)
    log.info("keyframe extract start: %s n=%d", video_path.name, n)

    # Получаем длительность
    rc, stdout, _ = await _run_ffmpeg(
        ["ffprobe", "-v", "error", "-show_entries", "format=duration",
         "-of", "default=noprint_wrappers=1:nokey=1", str(video_path)],
        timeout=30.0,
    )
    try:
        duration = float(stdout.decode().strip() or "0") if rc == 0 else 0.0
    except ValueError:
        duration = 0.0
    log.info("keyframe extract: %s duration=%.1fs", video_path.name, duration)

    frames: list[Path] = []
    if duration <= 0:
        out = frames_dir / "frame%02d.jpg"
        await _run_ffmpeg(
            ["ffmpeg", "-y", "-i", str(video_path),
             "-vf", "fps=1/10,scale=640:-1", "-frames:v", str(n), str(out)],
            timeout=60.0,
        )
    else:
        step = duration / (n + 1)
        for i in range(n):
            ts = step * (i + 1)
            out = frames_dir / f"frame_{i:02d}.jpg"
            rc2, _, _ = await _run_ffmpeg(
                ["ffmpeg", "-y", "-ss", str(ts), "-i", str(video_path),
                 "-vf", "scale=640:-1", "-frames:v", "1", "-q:v", "3", str(out)],
                timeout=30.0,
            )
            if rc2 == 0 and out.exists():
                frames.append(out)
    log.info("keyframe extract done: %s -> %d frames", video_path.name, len(frames))
    return frames


async def analyze_video(video_path: Path, describe_image_fn) -> str:
    """Анализировать видео: STT из аудио + описание 4 кейфреймов.

    describe_image_fn — async callable(base64_data: str, mime: str) -> str
    Возвращает итоговое текстовое описание.
    """
    parts: list[str] = []

    log.info("analyze_video start: %s", video_path.name)
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)

        # 1. STT
        audio_path = tmp / "audio.wav"
        ok = await _extract_video_audio_async(video_path, audio_path)
        if ok and audio_path.exists():
            log.info("video STT start: %s", video_path.name)
            transcript = await transcribe_audio(audio_path)
            log.info("video STT done: %s -> %r", video_path.name, transcript[:80] if transcript else "")
            if transcript and not transcript.startswith("("):
                parts.append(f"**Аудио (транскрипция):**\n{transcript}")
            else:
                log.info("video STT empty/failed: %s", transcript)
        else:
            log.info("no audio extracted from video %s", video_path.name)

        # 2. Keyframes → moondream
        frames_dir = tmp / "frames"
        frames = await _extract_keyframes_async(video_path, frames_dir, 4)
        if frames:
            frame_descriptions: list[str] = []
            for i, frame_path in enumerate(frames):
                try:
                    log.info("moondream frame %d/%d: %s", i + 1, len(frames), video_path.name)
                    img_bytes = frame_path.read_bytes()
                    b64 = base64.b64encode(img_bytes).decode()
                    desc = await describe_image_fn(b64, "image/jpeg")
                    log.info("moondream frame %d done: %r", i + 1, desc[:60])
                    frame_descriptions.append(f"Кадр {i + 1}: {desc}")
                except Exception as exc:  # noqa: BLE001
                    log.warning("frame %d describe failed: %s", i, exc)
            if frame_descriptions:
                parts.append("**Визуальный анализ кадров:**\n" + "\n".join(frame_descriptions))

    if not parts:
        return "(не удалось проанализировать видео)"
    return "\n\n".join(parts)


async def extract_text_from_file(path: Path) -> str | None:
    """Попытаться извлечь текст из файла по расширению.

    Возвращает строку с текстом или None если тип не поддерживается.
    """
    ext = path.suffix.lower()

    if ext == ".pdf":
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _extract_pdf, path)

    if ext in {".docx"}:
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _extract_docx, path)

    if ext in {".xlsx", ".xls"}:
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _extract_xlsx, path)

    if ext in TEXT_EXTENSIONS:
        return _extract_text(path)

    if ext in AUDIO_EXTENSIONS:
        return await transcribe_audio(path)

    # Видео: требует describe_image_fn — вернём None здесь,
    # вызов через analyze_video из tools.py напрямую
    if ext in VIDEO_EXTENSIONS:
        return None

    return None
